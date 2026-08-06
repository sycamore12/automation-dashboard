import streamlit as st
import pdfplumber
import openpyxl
from openpyxl.styles import PatternFill
import pandas as pd
import io
import os
import re

# Set Streamlit Page Configuration
st.set_page_config(
    page_title="Sistem Automasi BPJS Ketenagakerjaan",
    page_icon="📊",
    layout="wide"
)

# --- CUSTOM CSS FOR BPJS KETENAGAKERJAAN THEME ---
custom_css = """
<style>
    .bpjs-banner {
        height: 8px;
        width: 100%;
        background: linear-gradient(90deg, #008C44 33.3%, #F6EA00 33.3%, #F6EA00 66.6%, #005C9A 66.6%);
        margin-top: -10px;
        margin-bottom: 25px;
        border-radius: 4px;
    }
    div.stButton > button:first-child {
        background-color: #008C44 !important;
        color: white !important;
        border: none !important;
        font-weight: bold;
    }
    div.stButton > button:first-child:hover {
        background-color: #007036 !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
    }
    div.stDownloadButton > button:first-child {
        background-color: #005C9A !important;
        color: white !important;
        border: none !important;
        font-weight: bold;
    }
    div.stDownloadButton > button:first-child:hover {
        background-color: #004777 !important;
        box-shadow: 0 4px 8px rgba(0,0,0,0.1);
    }
</style>
"""
st.markdown(custom_css, unsafe_allow_html=True)
# --------------------------------------------------

# ==========================================
# SHARED UTILITIES
# ==========================================
MONTHS = [
    "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI", 
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER"
]

MONTH_MAP = {
    "01": "JANUARI", "02": "FEBRUARI", "03": "MARET", "04": "APRIL",
    "05": "MEI", "06": "JUNI", "07": "JULI", "08": "AGUSTUS",
    "09": "SEPTEMBER", "10": "OKTOBER", "11": "NOVEMBER", "12": "DESEMBER"
}

def get_month_from_filename(filename):
    upper_filename = filename.upper()
    for month in MONTHS:
        if month in upper_filename:
            return month
    return None

# ==========================================
# MONTHLY RECAP FUNCTIONS
# ==========================================
def extract_monthly_pdf_data(pdf_file_bytes):
    extracted_data = {}
    with pdfplumber.open(pdf_file_bytes) as pdf:
        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 7:
                        continue
                    kode_perisai = str(row[1]).strip() if row[1] else ""
                    if kode_perisai.startswith("AB"):
                        try:
                            val_jht_jkk_jkm = int(row[5]) if row[5] and str(row[5]).strip().isdigit() else 0
                            val_jkk_jkm = int(row[6]) if row[6] and str(row[6]).strip().isdigit() else 0
                            extracted_data[kode_perisai] = {
                                'JHT_JKK_JKM': val_jht_jkk_jkm,
                                'JKK_JKM': val_jkk_jkm
                            }
                        except ValueError:
                            continue
    return extracted_data

def process_monthly_excel_update(excel_bytes, all_pdf_data, log_container):
    wb = openpyxl.load_workbook(excel_bytes)
    sheet = wb.active
    total_updates = 0
    
    for month, data in all_pdf_data.items():
        month_col_start = None
        for col in range(1, sheet.max_column + 1):
            cell_value = str(sheet.cell(row=1, column=col).value).strip().upper()
            if cell_value == month:
                month_col_start = col
                break
                
        if not month_col_start:
            log_container.warning(f"⚠️ Month '{month}' was not found in Row 1 of the Excel template.")
            continue
            
        month_updates = 0
        for row in range(3, sheet.max_row + 1):
            excel_kode = str(sheet.cell(row=row, column=3).value).strip() if sheet.cell(row=row, column=3).value else ""
            if excel_kode in data:
                sheet.cell(row=row, column=month_col_start).value = data[excel_kode]['JHT_JKK_JKM']
                sheet.cell(row=row, column=month_col_start + 1).value = data[excel_kode]['JKK_JKM']
                month_updates += 1
                
        total_updates += month_updates
        log_container.success(f"✅ Updated {month_updates} rows for **{month}**.")

    red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
    for col in range(1, sheet.max_column + 1):
        header_value = str(sheet.cell(row=1, column=col).value).strip().upper()
        if header_value in MONTHS:
            for row in range(3, sheet.max_row + 1):
                kode_perisai_val = sheet.cell(row=row, column=3).value
                if kode_perisai_val is None or str(kode_perisai_val).strip() == "":
                    continue  
                
                cell_1 = sheet.cell(row=row, column=col)
                if cell_1.value is None or str(cell_1.value).strip() == "" or str(cell_1.value).strip() == "0":
                    cell_1.fill = red_fill
                    
                cell_2 = sheet.cell(row=row, column=col + 1)
                if cell_2.value is None or str(cell_2.value).strip() == "" or str(cell_2.value).strip() == "0":
                    cell_2.fill = red_fill
        
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    return output_stream, total_updates

# ==========================================
# DAILY RECAP FUNCTIONS
# ==========================================
def extract_daily_pdf_data(pdf_file_bytes):
    extracted_data = {}
    periode_string = "PERIODE TIDAK DITEMUKAN"
    
    with pdfplumber.open(pdf_file_bytes) as pdf:
        first_page_text = pdf.pages[0].extract_text()
        if first_page_text:
            match = re.search(r"\d{2}-\d{2}-\d{4}\s*s/d\s*\d{2}-\d{2}-\d{4}", first_page_text)
            if match:
                periode_string = match.group(0)

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 9: 
                        continue
                    
                    kode_perisai = str(row[1]).strip() if row[1] else ""
                    if kode_perisai.startswith("AB"):
                        try:
                            val_baru_1 = int(row[5]) if row[5] and str(row[5]).strip().isdigit() else 0
                            val_baru_2 = int(row[6]) if row[6] and str(row[6]).strip().isdigit() else 0
                            val_lanj_1 = int(row[7]) if row[7] and str(row[7]).strip().isdigit() else 0
                            val_lanj_2 = int(row[8]) if row[8] and str(row[8]).strip().isdigit() else 0
                            
                            extracted_data[kode_perisai] = {
                                'BARU_JHT': val_baru_1,
                                'BARU_JKK': val_baru_2,
                                'LANJ_JHT': val_lanj_1,
                                'LANJ_JKK': val_lanj_2
                            }
                        except ValueError:
                            continue
    return periode_string, extracted_data

def process_daily_excel_update(excel_bytes, extracted_data, periode_string, filename, log_container):
    wb = openpyxl.load_workbook(excel_bytes)
    sheet = wb.active
    
    sheet['E1'] = periode_string
    
    updates = 0
    for row in range(4, sheet.max_row + 1):
        excel_kode = str(sheet.cell(row=row, column=3).value).strip() if sheet.cell(row=row, column=3).value else ""
        
        if excel_kode in extracted_data:
            sheet.cell(row=row, column=5).value = extracted_data[excel_kode]['BARU_JHT']
            sheet.cell(row=row, column=6).value = extracted_data[excel_kode]['BARU_JKK']
            sheet.cell(row=row, column=7).value = extracted_data[excel_kode]['LANJ_JHT']
            sheet.cell(row=row, column=8).value = extracted_data[excel_kode]['LANJ_JKK']
            updates += 1
            
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    
    log_container.success(f"✅ Updated {updates} rows in `{filename}`.")
    return output_stream

# ==========================================
# TK AKTIF RECAP FUNCTIONS (NEW)
# ==========================================
def extract_tk_aktif_pdf_data(pdf_file_bytes):
    extracted_data = {}
    month_name = "BULAN TIDAK DITEMUKAN"
    
    with pdfplumber.open(pdf_file_bytes) as pdf:
        # Extract period to get the month
        first_page_text = pdf.pages[0].extract_text()
        if first_page_text:
            # Matches formats like "Periode 07-2026"
            match = re.search(r"(?i)Periode\s+(\d{2})-\d{4}", first_page_text)
            if match:
                month_num = match.group(1)
                month_name = MONTH_MAP.get(month_num, month_name)

        for page in pdf.pages:
            tables = page.extract_tables()
            for table in tables:
                if not table:
                    continue
                for row in table:
                    if not row or len(row) < 6: 
                        continue
                    
                    # Column index 2 is KODE PERISAI based on new image
                    kode_perisai = str(row[2]).strip() if row[2] else ""
                    
                    if kode_perisai.startswith("AB"):
                        try:
                            # Index 4: JUMLAH PENDAFTARAN (TK)
                            # Index 5: JUMLAH KEPS AKTIF (TK)
                            akuisisi = int(row[4]) if row[4] and str(row[4]).strip().isdigit() else 0
                            tk_aktif = int(row[5]) if row[5] and str(row[5]).strip().isdigit() else 0
                            
                            extracted_data[kode_perisai] = {
                                'AKUISISI': akuisisi,
                                'TK_AKTIF': tk_aktif
                            }
                        except ValueError:
                            continue
                            
    return month_name, extracted_data

def process_tk_aktif_excel_update(excel_bytes, extracted_data, month_name, filename, log_container):
    wb = openpyxl.load_workbook(excel_bytes)
    sheet = wb.active
    
    # Overwrite the BULAN cell (E1) with the extracted month name
    sheet['E1'] = month_name
    
    updates = 0
    # TK Aktif template data starts at Row 3
    for row in range(3, sheet.max_row + 1):
        excel_kode = str(sheet.cell(row=row, column=3).value).strip() if sheet.cell(row=row, column=3).value else ""
        
        if excel_kode in extracted_data:
            sheet.cell(row=row, column=5).value = extracted_data[excel_kode]['AKUISISI']
            sheet.cell(row=row, column=6).value = extracted_data[excel_kode]['TK_AKTIF']
            updates += 1
            
    output_stream = io.BytesIO()
    wb.save(output_stream)
    output_stream.seek(0)
    
    log_container.success(f"✅ Updated {updates} rows in `{filename}`.")
    return output_stream

# ==========================================
# MAIN UI
# ==========================================
st.title("Sistem Automasi Rekapitulasi Data")
st.markdown('<div class="bpjs-banner"></div>', unsafe_allow_html=True)

# Three tabs for the three different operations
tab1, tab2, tab3 = st.tabs(["📅 Monthly Recap", "📝 Daily Recap", "📈 TK Aktif Recap"])

# ------------------------------------------
# TAB 1: MONTHLY RECAP
# ------------------------------------------
with tab1:
    st.markdown("Upload your master Excel template and monthly PDF recaps below to process and merge data automatically.")
    
    st.subheader("📁 Step 1: Upload Files (Monthly)")
    col_m1, col_m2 = st.columns(2)
    with col_m1:
        monthly_excel_file = st.file_uploader("Upload HASIL_REKAP.xlsx", type=["xlsx"], key="monthly_excel")
    with col_m2:
        monthly_pdf_files = st.file_uploader("Upload PDF Recaps (Bulk)", type=["pdf"], accept_multiple_files=True, key="monthly_pdfs")

    st.divider()

    if st.button("🚀 Process Monthly Recap", type="primary", use_container_width=True, key="btn_monthly"):
        if not monthly_excel_file:
            st.error("Please upload the `HASIL_REKAP.xlsx` template file.")
        elif not monthly_pdf_files:
            st.error("Please upload at least one recap PDF file.")
        else:
            log_box = st.container()
            all_pdf_data = {}
            
            for pdf in monthly_pdf_files:
                target_month = get_month_from_filename(pdf.name)
                if not target_month:
                    log_box.error(f"Skipping `{pdf.name}`: Could not detect valid month in filename.")
                    continue
                    
                pdf_bytes = io.BytesIO(pdf.read())
                extracted = extract_monthly_pdf_data(pdf_bytes)
                
                if extracted:
                    all_pdf_data[target_month] = extracted
                    log_box.info(f"Extracted **{len(extracted)}** records from `{pdf.name}` for **{target_month}**.")
                else:
                    log_box.warning(f"No valid records found in `{pdf.name}`.")
                    
            if all_pdf_data:
                excel_bytes = io.BytesIO(monthly_excel_file.read())
                updated_excel_stream, total_updates = process_monthly_excel_update(excel_bytes, all_pdf_data, log_box)
                
                st.balloons()
                st.success(f"🎉 Process completed successfully! Total rows updated: **{total_updates}**")
                
                st.download_button(
                    label="📥 Download Updated HASIL_REKAP.xlsx",
                    data=updated_excel_stream,
                    file_name="HASIL_REKAP_UPDATED.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )

# ------------------------------------------
# TAB 2: DAILY RECAP
# ------------------------------------------
with tab2:
    st.markdown("Upload your daily PDF recap and the 3 regional Excel templates (Kecamatan/Wilayah) to process daily NIK Baru & Lanjutan.")
    
    st.subheader("📁 Step 1: Upload Files (Daily)")
    col_d1, col_d2 = st.columns(2)
    with col_d1:
        daily_excel_files = st.file_uploader("Upload Regional Excel Templates (Max 3)", type=["xlsx"], accept_multiple_files=True, key="daily_excels")
    with col_d2:
        daily_pdf_file = st.file_uploader("Upload Daily PDF Recap (Single)", type=["pdf"], accept_multiple_files=False, key="daily_pdf")

    st.divider()

    if st.button("🚀 Process Daily Recap", type="primary", use_container_width=True, key="btn_daily"):
        if not daily_excel_files or len(daily_excel_files) == 0:
            st.error("Please upload the regional Excel template files.")
        elif not daily_pdf_file:
            st.error("Please upload the daily recap PDF file.")
        else:
            log_box = st.container()
            
            pdf_bytes = io.BytesIO(daily_pdf_file.read())
            periode_str, extracted_daily_data = extract_daily_pdf_data(pdf_bytes)
            
            log_box.info(f"Detected Periode: **{periode_str}**")
            
            if extracted_daily_data:
                log_box.info(f"Extracted **{len(extracted_daily_data)}** records from `{daily_pdf_file.name}`.")
                
                st.success("🎉 Processing complete! Download your regional files below:")
                download_cols = st.columns(len(daily_excel_files))
                
                for index, excel_file in enumerate(daily_excel_files):
                    excel_bytes = io.BytesIO(excel_file.read())
                    updated_stream = process_daily_excel_update(
                        excel_bytes, 
                        extracted_daily_data, 
                        periode_str, 
                        excel_file.name, 
                        log_box
                    )
                    
                    new_filename = f"UPDATED_{excel_file.name}"
                    
                    with download_cols[index]:
                        st.download_button(
                            label=f"📥 {new_filename}",
                            data=updated_stream,
                            file_name=new_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_daily_{index}"
                        )
            else:
                st.error("Could not extract any valid data from the provided PDF.")

# ------------------------------------------
# TAB 3: TK AKTIF RECAP
# ------------------------------------------
with tab3:
    st.markdown("Upload your TK Aktif PDF recap and the 3 regional Excel templates (Kecamatan/Wilayah) to process data.")
    
    st.subheader("📁 Step 1: Upload Files (TK Aktif)")
    col_t1, col_t2 = st.columns(2)
    with col_t1:
        tk_aktif_excel_files = st.file_uploader("Upload Regional Excel Templates (Max 3)", type=["xlsx"], accept_multiple_files=True, key="tk_aktif_excels")
    with col_t2:
        tk_aktif_pdf_file = st.file_uploader("Upload TK Aktif PDF Recap (Single)", type=["pdf"], accept_multiple_files=False, key="tk_aktif_pdf")

    st.divider()

    if st.button("🚀 Process TK Aktif Recap", type="primary", use_container_width=True, key="btn_tk_aktif"):
        if not tk_aktif_excel_files or len(tk_aktif_excel_files) == 0:
            st.error("Please upload the regional Excel template files.")
        elif not tk_aktif_pdf_file:
            st.error("Please upload the TK Aktif recap PDF file.")
        else:
            log_box = st.container()
            
            pdf_bytes = io.BytesIO(tk_aktif_pdf_file.read())
            month_str, extracted_tk_aktif_data = extract_tk_aktif_pdf_data(pdf_bytes)
            
            log_box.info(f"Detected Bulan: **{month_str}**")
            
            if extracted_tk_aktif_data:
                log_box.info(f"Extracted **{len(extracted_tk_aktif_data)}** records from `{tk_aktif_pdf_file.name}`.")
                
                st.success("🎉 Processing complete! Download your regional files below:")
                download_cols = st.columns(len(tk_aktif_excel_files))
                
                for index, excel_file in enumerate(tk_aktif_excel_files):
                    excel_bytes = io.BytesIO(excel_file.read())
                    updated_stream = process_tk_aktif_excel_update(
                        excel_bytes, 
                        extracted_tk_aktif_data, 
                        month_str, 
                        excel_file.name, 
                        log_box
                    )
                    
                    new_filename = f"UPDATED_{excel_file.name}"
                    
                    with download_cols[index]:
                        st.download_button(
                            label=f"📥 {new_filename}",
                            data=updated_stream,
                            file_name=new_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key=f"dl_tk_aktif_{index}"
                        )
            else:
                st.error("Could not extract any valid data from the provided PDF.")
